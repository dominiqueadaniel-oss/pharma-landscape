"""
Parse the pharma competitive landscape spreadsheet into data.json.
Run once locally, or as part of the weekly refresh pipeline.
"""
import json
import sys
from pathlib import Path
from datetime import date

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"], check=True)
    import openpyxl

XLSX_PATH = Path(__file__).parent.parent / "data" / "source.xlsx"
OUTPUT_PATH = Path(__file__).parent.parent / "data" / "data.json"


def cell_val(ws, row, col):
    v = ws.cell(row, col).value
    return str(v).strip() if v is not None else ""


def find_header_row(ws, first_col_value):
    for r in range(1, min(12, ws.max_row + 1)):
        if ws.cell(r, 1).value == first_col_value:
            return r
    return None


def parse_rows(ws, header_row, key_map):
    headers = []
    for c in range(1, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        headers.append(str(h).strip().replace("\n", " ") if h else "")

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        raw = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if all(v is None for v in raw):
            continue
        if raw[0] is None:
            continue
        # Skip visual separator / section header rows (start with ── or similar)
        first = str(raw[0]).strip()
        if first.startswith("──") or first.startswith("--") or first.startswith("==="):
            continue
        item = {}
        for i, h in enumerate(headers):
            if h and i < len(raw):
                key = key_map.get(h, h)
                v = raw[i]
                item[key] = str(v).strip() if v is not None else ""
        rows.append(item)
    return rows


def parse_company_overview(ws):
    hr = find_header_row(ws, "Rank\n(Mkt Cap)")
    if hr is None:
        return []
    key_map = {
        "Rank\n(Mkt Cap)": "rank",
        "Company": "company",
        "Ticker": "ticker",
        "Market Cap\n(USD Bn, Apr 2026)": "market_cap_bn",
        "HQ Country": "hq",
        "Primary Immunology Focus": "immuno_focus",
        "Primary ID Focus": "id_focus",
        "2025 Immuno Revenue\n(USD Bn, reported)": "immuno_revenue",
        "2025 ID Revenue\n(USD Bn, reported)": "id_revenue",
        "Key LOE Risk (next 5yr)": "loe_risk",
        "Primary Verification Source": "source",
    }
    return parse_rows(ws, hr, key_map)


def parse_pipeline(ws, area):
    hr = find_header_row(ws, "Company")
    if hr is None:
        return []
    key_map = {
        "Company": "company",
        "Asset / Brand": "asset",
        "Mechanism of Action": "mechanism",
        "Mechanism / Modality": "mechanism",
        "Target": "target",
        "Pathogen / Disease Area": "target",
        "Indication(s)": "indication",
        "Indication / Population": "indication",
        "Phase / Status": "phase",
        "Key Trial Name(s)": "trial_names",
        "Key Trial / Study Name": "trial_names",
        "Primary Endpoint(s)": "endpoints",
        "Expected Data / PDUFA": "expected_data",
        "2025 Revenue\n(USD Bn)": "revenue_2025",
        "Peak Sales Consensus\n(USD Bn)": "peak_sales",
        "Peak Sales / Market\nPotential (USD Bn)": "peak_sales",
        "PoS Benchmark": "pos",
        "Risk Profile": "risk_profile",
        "Primary Source for Verification": "source",
    }
    rows = parse_rows(ws, hr, key_map)
    for r in rows:
        r["area"] = area
    return rows


def parse_marketed(ws):
    hr = find_header_row(ws, "Company")
    if hr is None:
        return []
    key_map = {
        "Company": "company",
        "Product (INN)": "product",
        "Therapeutic Area": "therapeutic_area",
        "Mechanism Class": "mechanism",
        "Approved Indications (US)": "indications",
        "US Approval Date": "approval_date",
        "2025 Revenue\n(USD Bn)": "revenue_2025",
        "US LOE Date\n(estimated)": "loe_date",
        "Active Biosimilar/Generic\nApplicants (US)": "biosimilar_applicants",
        "Competitive Threat Level": "threat_level",
        "Source": "source",
    }
    return parse_rows(ws, hr, key_map)


def parse_pos_table(ws):
    hr = find_header_row(ws, "Phase")
    if hr is None:
        return []
    key_map = {
        "Phase": "phase",
        "Overall Immuno PoS\n(Ph → Approval)": "pos_immuno",
        "Overall ID/Antiinfective PoS\n(Ph → Approval)": "pos_id",
        "Small Molecule vs. Biologic Modifier": "modifier",
        "Key Failure Modes": "failure_modes",
        "Due Diligence Flag": "dd_flag",
        "Source Reference": "source",
    }
    return parse_rows(ws, hr, key_map)


def parse_loe_watchlist(ws):
    hr = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == "Product":
            hr = r
            break
    if hr is None:
        return []
    key_map = {
        "Product": "product",
        "Company": "company",
        "Therapeutic Area": "therapeutic_area",
        "US LOE / Patent Expiry (est.)": "us_loe",
        "EU LOE (est.)": "eu_loe",
        "Active Biosimilar/Generic Applicants (US)": "biosimilar_applicants",
        "Projected Revenue Impact": "revenue_impact",
        "Priority Action": "action",
        "Source": "source",
    }
    return parse_rows(ws, hr, key_map)


def build_initial_editorial():
    return {
        "key_recent_events": [
            {
                "date": "2026-04",
                "headline": "AbbVie Skyrizi Reaches $18.1B in FY2025 Revenue",
                "detail": "Risankizumab (Skyrizi) delivered $18.1B in FY2025, ahead of consensus, driven by UC and CD label gains. Management reiterated a ~$21.5B 2026 guidance with a $25B+ 2028 target. Skyrizi + Rinvoq together now offset Humira's peak.",
                "companies": ["AbbVie"],
                "tags": ["IL-23", "Revenue Beat", "Immunology"]
            },
            {
                "date": "2025-06",
                "headline": "FDA Approves Gilead's Yeztugo (Lenacapavir) for HIV PrEP",
                "detail": "Lenacapavir's twice-yearly SC injection received FDA approval as Yeztugo for PrEP following PURPOSE 1 & 2 trials showing near-zero HIV infections vs. background. CHMP positive opinion received in EU. WHO prequalification granted October 2025 enabling LMIC access.",
                "companies": ["Gilead Sciences"],
                "tags": ["HIV", "PrEP", "FDA Approval", "Milestone"]
            },
            {
                "date": "2025-01",
                "headline": "Stelara US Biosimilars Launch — J&J LOE Clock Starts",
                "detail": "Multiple ustekinumab biosimilars (Wezlana/Amgen, Selarsdi/Teva-Alvotech, Yesintek, Pyzchiva) launched in January 2025. J&J faces a projected $5–7B global revenue erosion over the next 3–5 years, with Tremfya and nipocalimab as key offsets.",
                "companies": ["J&J"],
                "tags": ["Biosimilar", "LOE", "IL-12/23", "Stelara"]
            },
            {
                "date": "2025-03",
                "headline": "Dupixent Approved for Bullous Pemphigoid — Now 10+ Indications",
                "detail": "Dupilumab (Dupixent) gained US approval for bullous pemphigoid in 2025, its 10th+ approved indication. FY2025 revenue reached ~$21.0B (Sanofi-reported total). Biosimilar filing horizon is ~2030, giving Sanofi/Regeneron a substantial runway.",
                "companies": ["Sanofi", "Regeneron"],
                "tags": ["IL-4/13", "Dupixent", "Label Expansion", "Immunology"]
            },
            {
                "date": "2026-Q1",
                "headline": "AstraZeneca Beyfortus (Nirsevimab) Hits ~$1.9B in ID Revenue",
                "detail": "Co-marketed with Sanofi, nirsevimab for RSV prevention in infants delivered ~$1.9B in FY2025 ID revenue for AZN. ACIP recommendation and broad paediatric adoption drove uptake. RSV prevention is becoming a standard-of-care category.",
                "companies": ["AstraZeneca", "Sanofi"],
                "tags": ["RSV", "Vaccine", "Infectious Disease", "Revenue Beat"]
            }
        ],
        "implications": [
            {
                "theme": "IL-23 Class Now Defines Immunology Leadership",
                "body": "Skyrizi (AbbVie) and Tremfya (J&J) have established IL-23 p19 blockade as the dominant mechanism across psoriasis, IBD, and PsA. With combined revenues exceeding $21B and growing, IL-23 mAbs have displaced TNF inhibitors as the reference standard. Companies without a competitive IL-23 or next-gen program face structural share loss in core inflammatory markets.",
                "impact": "High",
                "companies_affected": ["Pfizer", "Novartis", "UCB"]
            },
            {
                "theme": "Long-Acting Injectable HIV Prevention Reshapes PrEP Market",
                "body": "FDA approval of Yeztugo (lenacapavir SC, twice-yearly) marks a paradigm shift from daily oral PrEP. Gilead's PURPOSE trial data (near-zero infections) sets a gold standard that oral PrEP cannot match on adherence-adjusted real-world efficacy. ViiV's Apretude (monthly IM) faces a superior competitor. The once-yearly formulation (PURPOSE 3, data ~2027) could be transformative — projecting 15B+ peak sales potential.",
                "impact": "Transformative",
                "companies_affected": ["ViiV/GSK", "Merck", "Gilead Sciences"]
            },
            {
                "theme": "Biosimilar Wave is Accelerating — LOE Cliff is Now",
                "body": "Humira (AbbVie), Stelara (J&J), and Actemra (Roche) are all mid-erosion. Cosentyx (Novartis) faces US biosimilar entry ~2027, Taltz (Lilly) ~2027, and Enbrel (Amgen) post-2029. Companies must demonstrate pipeline offset capability — AbbVie is the leading example with Skyrizi+Rinvoq. Those without a clear commercial successor face 15–30% revenue headwinds.",
                "impact": "High",
                "companies_affected": ["Novartis", "Amgen", "Pfizer", "Roche", "BMS"]
            },
            {
                "theme": "JAK Inhibitor Class Faces Regulatory Ceiling",
                "body": "FDA black box warning on JAK inhibitors (tofacitinib, baricitinib, upadacitinib) limits commercial penetration in less severe disease. Rinvoq (upadacitinib) is executing well in RA/IBD/AD despite the warning, but the ceiling on mild-moderate AD and new indications is constrained by prescriber hesitancy. Next-gen TYK2 inhibitors (e.g., deucravacitinib/Sotyktu) are being positioned as the safer alternative.",
                "impact": "Medium",
                "companies_affected": ["AbbVie", "Pfizer", "Eli Lilly", "BMS"]
            },
            {
                "theme": "FcRn Inhibitors: The Next Platform Battleground",
                "body": "Multiple companies (J&J/nipocalimab, UCB/rozanolixizumab, argenx/efgartigimod, Immunovant) are racing to establish FcRn blockade as a platform across IgG-driven diseases (MG, CIDP, warm hemolytic anemia, pemphigus). The addressable market is broad but fragmented. argenx (Vyvgart) has first-mover advantage; J&J's nipocalimab (Ph3) in pregnancy-related autoimmunity is a differentiated bet.",
                "impact": "Medium",
                "companies_affected": ["J&J", "UCB", "argenx", "Immunovant"]
            }
        ],
        "commentary": [
            {
                "company": "AbbVie",
                "headline": "Executing the Humira Succession Playbook — Better Than Expected",
                "body": "AbbVie's post-Humira transition is the clearest execution story in large-cap pharma. Skyrizi ($18.1B in FY2025, guidance $21.5B in 2026) and Rinvoq ($7.9B in FY2025, guided to $15B+ across all indications) together now nearly match Humira's peak. The next catalysts — Rinvoq approval in HS, SLE (Ph3 data 2026), vitiligo, and AA — extend the label footprint without requiring new molecular bet. Key risk: Rinvoq's JAK class Black Box remains a commercial ceiling; any new safety signal would be material. Lutikizumab (IL-1α/β) in HS is the swing option — if it works in combo with Skyrizi, it adds a meaningful new revenue layer.",
                "rating": "Strong Execution",
                "risk": "Low-Medium"
            },
            {
                "company": "Gilead Sciences",
                "headline": "HIV Platform Remains the Crown Jewel; PrEP Could Be the Next $10B Chapter",
                "body": "Gilead's HIV business is structurally dominant: Biktarvy ($12.4B in FY2025) is the global treatment standard through 2033 patent protection, and Yeztugo (lenacapavir PrEP, approved June 2025) opens a new prevention market that dwarfs any prior PrEP product. The PURPOSE 1 & 2 data is arguably the most compelling clinical dataset in HIV since integrase inhibitors. The once-yearly formulation (PURPOSE 3, data ~2027) is a speculative but potentially industry-defining development — 100M+ high-risk individuals globally who cannot adhere to monthly or even biannual injections. Key execution risk: LMIC access strategy (voluntary licensing to 120+ countries) compresses ASP but is essential to avoid regulatory and reputational barriers globally.",
                "rating": "Strong Execution",
                "risk": "Low"
            },
            {
                "company": "J&J (Janssen)",
                "headline": "Dual Biosimilar Headwind + Promising Offset Portfolio — a Transition Year",
                "body": "J&J faces simultaneous Stelara erosion (biosimilars launched Jan 2025, $5–7B global risk over 5 years) while ramping Tremfya in IBD and advancing nipocalimab (FcRn) into Ph3. The Tremfya UC/CD data and nipocalimab in gestational alloimmune disease (Ph3) are the two pipeline catalysts most likely to reset the top-line growth narrative. Darzalex continues to be a significant immuno-oncology revenue contributor. Risk: Stelara erosion is faster than many models assumed; if Tremfya UC takes longer to ramp, J&J's immunology segment will show a revenue dip in 2025–2026 before recovering. The nipocalimab program in SLE and autoimmune indications is the longer-dated option value play.",
                "rating": "Transitional — Monitoring",
                "risk": "Medium"
            },
            {
                "company": "Sanofi / Regeneron",
                "headline": "Dupixent Franchise Remains the Most Durable Moat in Immunology",
                "body": "Dupixent ($21.0B in FY2025) has 10+ approved indications and a SLE Ph3 underway. Its multi-indication breadth creates a structural moat — biosimilars would need to replicate regulatory dossiers across dozens of indications and formulations. The $25B+ 2030 consensus implies continued penetration in COPD (a large market with significant unmet need), CSU (refractory patients), and CTCL. Key risk: the dupilumab biosimilar horizon (~2030) is approaching faster than the market prices in — first BPCIA filings likely ~2027. Sanofi's pipeline diversification (PCSK9, CD38, IL-33) also matters but is secondary to Dupixent execution in the near term.",
                "rating": "Strong Execution",
                "risk": "Low"
            },
            {
                "company": "AstraZeneca",
                "headline": "Respiratory Immunology Franchise Maturing; RSV a Bright Spot",
                "body": "AstraZeneca's immunology portfolio (Fasenra/IL-5, Tezspire/TSLP, Saphnelo/IFNα) is mature and faces biosimilar headwinds for Fasenra (~2027+). Tezspire's differentiation (TSLP, broader mechanism) is the key near-term growth driver. Saphnelo (SLE) has been a modest underperformer vs. peak expectations. The RSV story (Beyfortus ~$1.9B in ID, co-marketed with Sanofi) is a genuine bright spot — nirsevimab is now standard-of-care in infant RSV prevention. Pipeline adjacencies (COPD inflammation, eosinophilic diseases) are medium-term catalysts but face a competitive field.",
                "rating": "Stable — Monitoring",
                "risk": "Medium"
            }
        ],
        "last_generated": str(date.today())
    }


def main(xlsx_path=None):
    if xlsx_path is None:
        xlsx_path = XLSX_PATH

    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        print(f"ERROR: File not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path)

    companies = parse_company_overview(wb["COMPANY OVERVIEW"])
    immuno_pipeline = parse_pipeline(wb["IMMUNOLOGY PIPELINE"], "Immunology")
    id_pipeline = parse_pipeline(wb["ID PIPELINE"], "Infectious Disease")
    marketed = parse_marketed(wb["MARKETED PRODUCTS"])
    pos_table = parse_pos_table(wb["RISK & VALUE MATRIX"])
    loe_watchlist = parse_loe_watchlist(wb["RISK & VALUE MATRIX"])
    editorial = build_initial_editorial()

    data = {
        "meta": {
            "title": "Pharma & Biotech Competitive Landscape: Immunology & Infectious Disease",
            "subtitle": "Top 20 Companies by Market Cap | Pipeline + Marketed Products | Phase, Value & Risk Profile",
            "last_updated": str(date.today()),
            "data_as_of": "Q1 2026",
            "refresh_cadence": "Weekly",
        },
        "companies": companies,
        "pipeline": immuno_pipeline + id_pipeline,
        "marketed_products": marketed,
        "pos_benchmarks": pos_table,
        "loe_watchlist": loe_watchlist,
        "editorial": editorial,
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"Written: {OUTPUT_PATH}")
    print(f"  Companies: {len(companies)}")
    print(f"  Pipeline assets: {len(immuno_pipeline) + len(id_pipeline)} ({len(immuno_pipeline)} immuno, {len(id_pipeline)} ID)")
    print(f"  Marketed products: {len(marketed)}")
    print(f"  LOE watchlist: {len(loe_watchlist)}")


if __name__ == "__main__":
    xlsx = sys.argv[1] if len(sys.argv) > 1 else None
    main(xlsx)
